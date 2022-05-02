from encodings import utf_8
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(['번호', '국어', '영어'])

for i in range(1, 11):
    ws.append([i,randint(0,100), randint(0,100)])
"""
col_b = ws["B"] #B 컬럼만을 축출하여 가져옴
col_range = ws["b:c"]
row_title = ws[1]
for cell in col_b:
    print(cell.value)

for cols in col_range:
    for cell in cols:
        print(cell.value)

for cell in row_title:
    print(cell.value)

for cell in ws[2:6]:
    for rows in cell:
        print(rows.value, end=" ")
    print()


row_range = ws[2:ws.max_row]

for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

print(tuple(ws.rows))

for row in ws.iter_rows:
    print(row[2].value)
"""
for row in ws.iter_rows(min_row=2, max_row= 11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)
wb.save("sample3.xlsx")