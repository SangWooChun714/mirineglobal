from openpyxl import load_workbook

wb = load_workbook("sample3.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80 :
        print(row[0].value, "참잘했어요")
        
for row in ws.iter_rows(max_row=1):
    for cell in row:
        #print(cell.value, end=" ")
        if cell.value == "수학":
            cell.value = "국어"

wb.save("sample3.xlsx")