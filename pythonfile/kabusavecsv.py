from bs4 import BeautifulSoup
from datetime import date
from drawfin import draw
import sys, csv, requests, datetime, re, ssl
from openpyxl import load_workbook
from kabuconfig import logger
from elasticsearch import Elasticsearch
import pandas as pd


holyday_list = ["0101", "0301", "0505", "0606", "0815", "1003", "1009", "1225"] #公休日list
todays = str(date.today().strftime("%Y.%m.%d")) # 今日の日付を保存
__headers = {"user-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36"}#heaaderを指定

#探すtableを確認する関数
def table_number(url) :
    logger.info("start tablenumber check")
    try:
        res1 = requests.get(url, headers=__headers) #資料を持ってくる注所を設定
        searchday = BeautifulSoup(res1.text, "lxml")
        searchtablenum = searchday.find("td", attrs={"class":"pgRR"}) # 探すpageの終わりを確認
        if not searchtablenum:
            searchtablenum = searchday.find("table", attrs={"class":"Nnavi"}).find("td",{"class":"on"})
            tablenum = searchtablenum.a.get("href").rsplit("=")[2]
        else :
            tablenum = searchtablenum.a.get("href").rsplit("=")[2]
        logger.info("url : "+url+", tablenum : "+tablenum)
    except Exception as e:
        print(e)
        logger.warning(e)
    logger.info("end tablenumber check")
    return tablenum

#株価をcrawlingする関数
#totalで架空した資料を保存
#tablenumで持ってくるpageを決定
#dayは今処理しているひ、daysは入力された日
#bodyはelasticsearchのindexに入れる
def SearchKabuKa(kabunames, code, days, filename):
    total = [] # 会社の名前, 日付, 株価, 取引量を保存するlist
    breaker = True # 無限文を止める変数
    url = "https://finance.naver.com/item/sise_day.naver?code="+code # 株価を持ってくる住所

    ctx = ssl.create_default_context()
    #ctx.load_verify_locations("C:/Users/cjstk/Desktop/python/elasticsearch-8.1.0/config/certs/http_ca.crt")
    #ctx.load_verify_locations("C:/Users/cjstk/Desktop/elasticsearch-8.1.0/cert1/http_ca.crt")
    ctx.load_verify_locations("C:/Users/mg-e1/Desktop/elasticsearch-8.1.0/cert1/http_ca.crt")
    es = Elasticsearch("https://localhost:9200", basic_auth=("elastic","sw1594311"), ssl_context=ctx)
    #es = Elasticsearch("http://localhost:9200/")
    index = "kabusearch"
    n = 1
    
    logger.info("start kabusearch")
    tablenum = table_number(url)
    try:
        f = open(filename, "w", encoding="utf-8-sig", newline="") # csvファイルの作成
        logger.info("create csv file")
        writer = csv.writer(f)
        tittle = ["name","code", "day", "price", "qunt"] #　一行をインデックスで作成
        writer.writerow(tittle)

        for i in range(1, int(tablenum)+1): # 最初から最終まで読む
            items2 = url+"&page="+str(i) # 住所とpage番号で接続
            res1 = requests.get(items2, headers=__headers)
            searchday = BeautifulSoup(res1.text, "lxml")
            table_rows = searchday.find("table", attrs={"class":"type2"}).find_all("tr") # tableから資料を持ってくる

            for row in table_rows:
                colums = row.find_all("td")

                if len(colums) <= 1:
                    continue

                data = [column.get_text().strip() for column in colums] # 必要の数値を持ってくる
                day = data[0] # 日付
                price = data[1] # 株価
                qunt = data[6] # 取引量
                if todays == day: #　今日の資料を除く
                    continue
                if day == days[0] or day == days[1]: # 指定した日まで保存
                    breaker = False

                    total = [kabunames, code, day, price, qunt]

                    body = {
                        "kabuNm" : kabunames,
                        "code" : code,
                        "date" : day,
                        "kabuka" : price,
                        "volume" : qunt
                    }
                    es.index(index=index, id=n, document=body)

                    writer.writerow(total)
                    break
                
                total = [kabunames, code, day, price, qunt]
                body = {
                        "kabuNm" : kabunames,
                        "code" : code,
                        "date" : day,
                        "kabuka" : price,
                        "volume" : qunt
                }
                es.index(index=index, id=n, document=body)
                n += 1

                writer.writerow(total) # csvファイルに一行づつ作成
                
            if breaker == True:
                continue
            else : 
                print("작업종료")
                logger.info("end csv saved")

                break

        f.close()

    except Exception as e:
        print(e)
        logger.warning(e)
    
#名前と日付が正常入力したかを確認する関数
#name_compilerで会社の名前の文字を確認する
#date_compilerで日の確認する
def checking(kabunames, days):
    logger.info("start checking")
    name_compiler = any(specialtext in kabunames for specialtext in "!@#$%^*()[]-=+_~`;'/?.<>, \t \n")
    print(name_compiler)

    if name_compiler == True :
        logger.warning("companyname error. input kabu : "+kabunames)
        print("주식이름 이상, 입력받은 주식이름 : "+kabunames)
        sys.exit()
    
    date_compiler = re.compile(r"([12]\d{3}).(0\d|1[0-2]).([0-2]\d|3[01])$")
    date_check = date_compiler.match(days)
    print(date_check)
    if not date_check:
        logger.warning("date error. input dayte : "+days)
        print("날짜를 정확히 입력해주세요. 입력받은 날짜 : "+days)
        sys.exit()
    logger.info("end checking")

#入力された日が平日か休日かを確認する関数
#yearsは年、monthsは月、daysは日を保存
#weekenddayは入力した日の曜日を数字で保存
#temp_dateは確認する日を仮に保存
#fine_daysは確定した日を保存
def FindDays(days) :
    fine_days = ["", ""]
    logger.info("start finddays")
    list_date = days.split(".")
    str_monday = list_date[1]+list_date[2] # 公休日を確認するための変数
    list_date = [int(list_date[0]), int(list_date[1]), int(list_date[2])]

    temp_date = datetime.date(list_date[0], list_date[1], list_date[2])
    weekendday = temp_date.weekday()
    
    if weekendday >= 5 or str_monday in holyday_list: # 休日ならその前の日を指定するためのif ex) 休日が金曜日の場合は木曜日を指定, 週末の場合金曜日を指定
        temp_date1 = str(temp_date - datetime.timedelta(max(1, (weekendday+6)%7-3))).split("-")
        temp_date2 = str(temp_date - datetime.timedelta(max(2, (weekendday+6)%7-2))).split("-")
        fine_days[0] = temp_date1[0]+"."+temp_date1[1]+"."+temp_date1[2]
        fine_days[1] = temp_date2[0]+"."+temp_date2[1]+"."+temp_date2[2]
    else : 
        temp_date = str(temp_date).split("-")
        fine_days[0] = (temp_date[0]+"."+temp_date[1]+"."+temp_date[2]) #　最終指定日を指定
    logger.info("end finddays : "+fine_days[0]+", "+fine_days[1])
    return fine_days

# main 関数
#argｖで　会社の名前と日付をもらう
#会社の名前は　kabuname　で保存
#日付は　inputdayｓ　で保存
if __name__ == "__main__":
    logger.info("program start")
    args = sys.argv
    logger.info("kabuname : "+args[1]+" day : "+args[2])
    print(args)
    kabunames = args[1] #　入力した会社の名前を保存
    inputdays = args[2] # 入力した日付を保存

    checking(kabunames, inputdays)#名前と日付の正規検索

    filename = kabunames + ".csv" # csvファイルの名前設定

    search_date = FindDays(inputdays)#休日と平日の確認
    
    wb = load_workbook("상장법인목록.xlsx") #会社のコードを確認のためファイルを読む
    ws = wb.active
    for row in ws.iter_rows(min_row=2): #　ファイルを読んで会社のコードと名前と日付をcrawling関数に入れる
        if row[0].value == kabunames:
            code = row[1].value
            wb.close()
            SearchKabuKa(kabunames, code, search_date, filename)#crawling関数
            break
    draws = draw()
    draws.drawline(filename)# 資料をもとにgraphを書く
    logger.info("program end")
    #drawfin.drawline(filename) 


